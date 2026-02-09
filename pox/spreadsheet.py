"""Spreadsheet generation for pox."""

from pathlib import Path
from typing import Any

from openpyxl.packaging.custom import StringProperty
from openpyxl.styles import PatternFill
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from . import __version__
from .datastructures import Message, SpreadsheetContext
from .styles import (
    EMPTY_BORDER,
    EMPTY_BORDER_BOTTOM,
    EMPTY_BORDER_SINGLE,
    EMPTY_BORDER_TOP,
    ROW_STRIPE,
)
from .styles import (
    SpreadsheetStyles as S,
)


def apply_styles(ws: Worksheet, row_idx: int, row_data: list | tuple) -> None:
    """Apply styles from (value, style) tuples to the cells of the given row."""
    for col_idx, cell_data in enumerate(row_data, start=1):
        if not isinstance(cell_data, tuple) or len(cell_data) < 2:  # pragma: no cover
            continue
        style = cell_data[1]
        if style is None:
            continue
        cell = ws.cell(row=row_idx, column=col_idx)
        style_dict = style.value
        for attr, value in style_dict.items():
            setattr(cell, attr, value)


def apply_empty_borders(ws: Worksheet, translation_col: int, num_cols: int) -> None:
    """
    Apply thick black borders around contiguous groups of empty translation cells.

    Scans translation columns and draws a thick black border box around each
    contiguous vertical run of empty (MSG_STR_EMPTY) cells.
    """
    # header is row 1, data starts at row 2
    data_start = 2
    data_end = ws.max_row

    for col in range(translation_col, translation_col + num_cols):
        # Collect row indices that have the empty style (yellow fill)
        empty_rows = []
        for row in range(data_start, data_end + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == "FFFFF2CC":
                empty_rows.append(row)

        # Group contiguous rows and apply borders
        if not empty_rows:
            continue

        groups = []
        start = empty_rows[0]
        prev = empty_rows[0]
        for r in empty_rows[1:]:
            if r == prev + 1:
                prev = r
            else:
                groups.append((start, prev))
                start = r
                prev = r
        groups.append((start, prev))

        for group_start, group_end in groups:
            if group_start == group_end:
                ws.cell(row=group_start, column=col).border = EMPTY_BORDER_SINGLE
            else:
                ws.cell(row=group_start, column=col).border = EMPTY_BORDER_TOP
                for r in range(group_start + 1, group_end):
                    ws.cell(row=r, column=col).border = EMPTY_BORDER
                ws.cell(row=group_end, column=col).border = EMPTY_BORDER_BOTTOM


def apply_row_stripes(ws: Worksheet, data_start: int, data_end: int) -> None:
    """Apply alternating row stripes to data rows for readability."""
    stripe_fill = PatternFill(fill_type=FILL_SOLID, fgColor=ROW_STRIPE)
    for row in range(data_start, data_end + 1):
        if (row - data_start) % 2 == 1:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                # Only apply stripe if the cell doesn't already have a fill
                if (
                    not cell.fill
                    or not cell.fill.fgColor
                    or cell.fill.fgColor.rgb == "00000000"
                ):
                    cell.fill = stripe_fill


class SpreadsheetGenerator:
    """Generate Excel spreadsheets from translation data."""

    outdir: Path

    def __init__(self, outdir: Path) -> None:
        """Initialize generator with output directory."""
        self.outdir = outdir

    def get_tabledata(
        self,
        messages: list[Message],
        context: SpreadsheetContext,
    ) -> list[Any]:
        """
        Generate a matrix array of messages for a table.

        id  Context     Singular Form   Translation
         1              Hello World     Hallo Welt
         2  Keep short  Sausage         Wurst
         3  obsolete    ~~Nudel~~       ~~Noodle~~

        Attach a format to each cell.
        """
        plural_hints = context.plural_form_hints or {}
        num_plurals = len(plural_hints)
        has_plurals = num_plurals >= 2 and any(m.is_plural for m in messages)

        # Extra columns beyond the base "Translation" column (which holds form 0)
        extra_cols = num_plurals - 1 if has_plurals else 0

        base_translation_label = "Translation"
        if has_plurals:
            hint = plural_hints.get(0, "Singular")
            base_translation_label = f"Translation ({hint})"

        header = [
            ("id", S.HEADER),
            ("Context", S.HEADER),
            ("Singular Form", S.HEADER),
            (base_translation_label, S.HEADER),
        ]

        # Add extra plural translation headers (forms 1, 2, ...)
        if has_plurals:
            for i in range(1, num_plurals):
                hint = plural_hints.get(i, f"Plural Form {i + 1}")
                header.append((f"Translation ({hint})", S.HEADER))

        data = []
        for i, m in enumerate(messages):
            msg_str_style = S.MSG_ID
            if m.obsolete:
                msg_str_style = S.MSG_STR_OBSOLETE
            elif not m.is_plural and m.translation.msgstr == "":
                msg_str_style = S.MSG_STR_EMPTY

            if m.is_plural:
                # For plural messages, the first translation is form 0 (singular),
                # the second is form 1 (plural), and any additional go into extra cols.
                msgstr_dict = m.translation.msgstr

                # Check if any plural form is empty
                if any(v == "" for v in msgstr_dict.values()):
                    msg_str_style = S.MSG_STR_EMPTY

                row = [
                    (i + 1, S.ID),
                    ("obsolete" if m.obsolete else m.context, S.CONTEXT),
                    (m.translation.msgid, S.MSG_ID),
                    (msgstr_dict.get(0, ""), msg_str_style),
                ]
                row.extend(
                    (msgstr_dict.get(pi, ""), msg_str_style)
                    for pi in range(1, num_plurals)
                )

                data.append(row)
            else:
                data.append(
                    [
                        (i + 1, S.ID),
                        ("obsolete" if m.obsolete else m.context, S.CONTEXT),
                        (m.translation.msgid, S.MSG_ID),
                        (m.translation.msgstr, msg_str_style),
                    ]
                    + [("", None)] * extra_cols,
                )

        return [
            header,
            *data,
        ]

    def generate(self, filename: str, context: SpreadsheetContext) -> Path:
        """Generate the spreadsheet file."""
        wb = Workbook()

        # Attach the language as a custom property
        props = [
            StringProperty(name="Language", value=context.language),
            StringProperty(name="PoxConvert", value=__version__),
        ]

        for p in props:
            wb.custom_doc_props.append(p)

        # Delete the auto generated "Sheet" worksheet
        del wb[wb.sheetnames[0]]

        # Create two sheets, "Translations" to hold the actual translations
        # and "Metadata" to store some extra data not needed for translation,
        # but might be useful upon parsing.
        ws: Worksheet = wb.create_sheet(f"Translations ({context.language})", index=0)
        wm: Worksheet = wb.create_sheet("Metadata", index=1)

        # Remove all gridlines right away
        ws.sheet_view.showGridLines = False
        wm.sheet_view.showGridLines = False

        # Generate the main spreadsheet for translations
        data = self.get_tabledata(
            messages=context.messages,
            context=context,
        )

        for row_idx, row in enumerate(data, start=1):
            ws.append([cell[0] for cell in row])
            apply_styles(ws, row_idx, row)

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Set column widths
        ws.column_dimensions["A"].width = 6  # id
        ws.column_dimensions["B"].width = 20  # Context
        ws.column_dimensions["C"].width = 55  # Singular Form
        ws.column_dimensions["D"].width = 55  # Translation
        # Extra plural columns
        for i in range(4, ws.max_column):
            col_letter = chr(ord("A") + i)
            ws.column_dimensions[col_letter].width = 55

        # Apply alternating row stripes (data starts at row 2)
        if ws.max_row > 1:
            apply_row_stripes(ws, data_start=2, data_end=ws.max_row)

        # Apply thick black borders around empty translation columns
        num_translation_cols = 1
        plural_hints = context.plural_form_hints or {}
        num_plurals = len(plural_hints)
        has_plurals = num_plurals >= 2 and any(m.is_plural for m in context.messages)
        if has_plurals:
            num_translation_cols = num_plurals
        apply_empty_borders(ws, translation_col=4, num_cols=num_translation_cols)

        # Write out the Excel spreadsheet and return its generated filename.
        filename = self.outdir / filename.format(
            lang=context.language,
            date=context.created.strftime("%Y-%m-%d"),
        )
        wb.save(filename)
        return filename
