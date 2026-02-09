"""Tests for Excel-to-PO conversion (xop-convert) and roundtrip."""

import argparse
from pathlib import Path

import polib
import pytest
from openpyxl import Workbook
from openpyxl.packaging.custom import StringProperty

from pox.base import ConverterError
from pox.cli.xop_convert import Excel2PoConverter


def test_singular_roundtrip(tmp_path: Path, singular_xlsx: Path) -> None:
    """Singular entries survive a PO-to-Excel-to-PO roundtrip."""
    opts = argparse.Namespace(
        xlsx_file=[str(singular_xlsx)],
        outdir=tmp_path,
        filename="{lang}_roundtrip.po",
    )
    converter = Excel2PoConverter(options=opts)
    converter.run()

    result_po = polib.pofile(str(tmp_path / "de_roundtrip.po"))
    msgids = [e.msgid for e in result_po]
    assert "Hello" in msgids
    assert "Goodbye" in msgids
    assert "Open" in msgids


def test_singular_translation_preserved(tmp_path: Path, singular_xlsx: Path) -> None:
    """Singular translation values are preserved through roundtrip."""
    opts = argparse.Namespace(
        xlsx_file=[str(singular_xlsx)],
        outdir=tmp_path,
        filename="{lang}.po",
    )
    converter = Excel2PoConverter(options=opts)
    converter.run()

    result_po = polib.pofile(str(tmp_path / "de.po"))
    entries = {e.msgid: e for e in result_po}
    assert entries["Hello"].msgstr == "Hallo"


def test_language_metadata_preserved(tmp_path: Path, singular_xlsx: Path) -> None:
    """Language metadata is preserved through roundtrip."""
    opts = argparse.Namespace(
        xlsx_file=[str(singular_xlsx)],
        outdir=tmp_path,
        filename="{lang}.po",
    )
    converter = Excel2PoConverter(options=opts)
    converter.run()

    result_po = polib.pofile(str(tmp_path / "de.po"))
    assert result_po.metadata["Language"] == "de"


def test_plural_roundtrip(tmp_path: Path, plurals_xlsx: Path) -> None:
    """Plural entries survive a PO-to-Excel-to-PO roundtrip."""
    opts = argparse.Namespace(
        xlsx_file=[str(plurals_xlsx)],
        outdir=tmp_path,
        filename="{lang}.po",
    )
    converter = Excel2PoConverter(options=opts)
    converter.run()

    result_po = polib.pofile(str(tmp_path / "de.po"))
    plural_entries = [e for e in result_po if e.msgid_plural]
    assert len(plural_entries) >= 1
    entry = plural_entries[0]
    assert entry.msgstr_plural[0] == "%d Gegenstand"
    assert entry.msgstr_plural[1] == "%d Gegenstaende"


def test_context_roundtrip(tmp_path: Path, singular_xlsx: Path) -> None:
    """Message context (msgctxt) is preserved through roundtrip."""
    opts = argparse.Namespace(
        xlsx_file=[str(singular_xlsx)],
        outdir=tmp_path,
        filename="{lang}.po",
    )
    converter = Excel2PoConverter(options=opts)
    converter.run()

    result_po = polib.pofile(str(tmp_path / "de.po"))
    entries = {e.msgid: e for e in result_po}
    assert entries["Open"].msgctxt == "adjective"


def test_bad_outdir_fails(tmp_path: Path) -> None:
    """Non-existent output directory raises ConverterError."""
    opts = argparse.Namespace(
        xlsx_file=[],
        outdir=tmp_path / "nonexistent",
        filename="{lang}.po",
    )
    with pytest.raises(ConverterError):
        Excel2PoConverter(options=opts)


def test_no_translations_sheet_fails(tmp_path: Path) -> None:
    """Missing Translations sheet raises ConverterError."""
    xlsx_path = tmp_path / "bad.xlsx"
    wb = Workbook()
    wb.worksheets[0].title = "NotTranslations"
    wb.save(str(xlsx_path))
    wb.close()

    opts = argparse.Namespace(
        xlsx_file=[str(xlsx_path)],
        outdir=tmp_path,
        filename="{lang}.po",
    )
    converter = Excel2PoConverter(options=opts)
    with pytest.raises(ConverterError):
        converter.run()


def test_not_enough_rows_fails(tmp_path: Path) -> None:
    """Sheet with insufficient rows raises ConverterError."""
    xlsx_path = tmp_path / "empty.xlsx"
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = "Translations (de)"
    # Only one row (header), no data — but we need < 2 rows total
    # So leave the sheet completely empty
    wb.save(str(xlsx_path))
    wb.close()

    opts = argparse.Namespace(
        xlsx_file=[str(xlsx_path)],
        outdir=tmp_path,
        filename="{lang}.po",
    )
    converter = Excel2PoConverter(options=opts)
    with pytest.raises(ConverterError):
        converter.run()


def test_no_language_property_uses_filename(tmp_path: Path) -> None:
    """Missing Language property falls back to filename extraction."""
    xlsx_path = tmp_path / "translations_fr.xlsx"
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = "Translations"
    ws.append(["id", "Context", "Singular Form", "Translation"])
    ws.append([1, None, "Hello", "Bonjour"])
    wb.save(str(xlsx_path))
    wb.close()

    opts = argparse.Namespace(
        xlsx_file=[str(xlsx_path)],
        outdir=tmp_path,
        filename="{lang}.po",
    )
    converter = Excel2PoConverter(options=opts)
    converter.run()

    result_po = polib.pofile(str(tmp_path / "fr.po"))
    assert result_po.metadata["Language"] == "fr"


def test_empty_rows_skipped(tmp_path: Path) -> None:
    """Empty/None rows in the xlsx data are skipped."""
    xlsx_path = tmp_path / "gaps.xlsx"
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = "Translations (de)"
    ws.append(["id", "Context", "Singular Form", "Translation"])
    ws.append([1, None, "Hello", "Hallo"])
    ws.append([None, None, None, None])  # empty row
    ws.append(["", None, None, None])  # blank id row
    ws.append([2, None, "World", "Welt"])
    wb.custom_doc_props.append(StringProperty(name="Language", value="de"))
    wb.save(str(xlsx_path))
    wb.close()

    opts = argparse.Namespace(
        xlsx_file=[str(xlsx_path)],
        outdir=tmp_path,
        filename="{lang}.po",
    )
    converter = Excel2PoConverter(options=opts)
    converter.run()

    result_po = polib.pofile(str(tmp_path / "de.po"))
    msgids = [e.msgid for e in result_po]
    assert "Hello" in msgids
    assert "World" in msgids
    assert len(result_po) == 2


def test_obsolete_entry_roundtrip(tmp_path: Path, singular_xlsx: Path) -> None:
    """Obsolete entries have context set to None (not 'obsolete' string)."""
    opts = argparse.Namespace(
        xlsx_file=[str(singular_xlsx)],
        outdir=tmp_path,
        filename="{lang}.po",
    )
    converter = Excel2PoConverter(options=opts)
    converter.run()

    result_po = polib.pofile(str(tmp_path / "de.po"))
    # The obsolete marker in context column should not leak as msgctxt
    for entry in result_po:
        assert entry.msgctxt != "obsolete"
