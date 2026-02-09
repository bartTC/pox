"""Convert Excel to PO files."""

import argparse
import sys
from pathlib import Path

import polib
from openpyxl import load_workbook

from pox.base import ArgumentFormatter, BaseConverter
from pox.utils import dedent


class Excel2PoConverter(BaseConverter):
    """Convert Excel spreadsheets back to PO files."""

    options: argparse.Namespace

    def __init__(self, options: argparse.Namespace) -> None:
        """Initialize converter and validate output directory."""
        self.options = options

        if options.outdir and not options.outdir.is_dir():
            self.fail(f'The output directory "{options.outdir}" does not exist.')

    def run(self) -> None:
        """Convert all provided xlsx files to PO format."""
        for xlsx_path in self.options.xlsx_file:
            self.convert_xlsx(Path(xlsx_path))

    def convert_xlsx(self, path: Path) -> None:  # noqa: C901
        """Parse an xlsx file and write a corresponding PO file."""
        wb = load_workbook(str(path), read_only=True, data_only=True)

        # Read custom properties for metadata
        language = None
        for prop in wb.custom_doc_props:
            if prop.name == "Language":
                language = prop.value

        if not language:
            self.warning(f'No "Language" property found in "{path}". Using filename.')
            language = path.stem.split("_", 1)[-1] if "_" in path.stem else "unknown"

        # Find the translations sheet
        ws = None
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("Translations"):
                ws = wb[sheet_name]
                break

        if ws is None:
            self.fail(f'No "Translations" sheet found in "{path}".')

        # Parse rows into polib entries
        po = polib.POFile()
        po.metadata = {
            "Content-Type": "text/plain; charset=UTF-8",
            "Language": language,
        }

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            self.fail(f'Not enough rows in "{path}".')

        # Row 0 = header, rows 1+ = data
        header_row = rows[0]

        # Determine number of translation columns
        # Columns: id, Context, Singular Form, Translation, [extra plural cols...]
        num_translation_cols = sum(
            1 for h in header_row[3:] if h and str(h).startswith("Translation")
        )

        data_rows = rows[1:]  # Skip header

        for row in data_rows:
            if not row or row[0] is None or row[0] == "":
                continue

            context = row[1] if row[1] and row[1] != "obsolete" else None
            obsolete = row[1] == "obsolete"
            msgid = row[2] or ""

            if num_translation_cols > 1:
                # Plural entry
                msgstr_plural = {}
                for pi in range(num_translation_cols):
                    col_idx = 3 + pi
                    val = row[col_idx] if col_idx < len(row) else ""
                    msgstr_plural[pi] = val or ""

                entry = polib.POEntry(
                    msgid=msgid,
                    msgid_plural=msgid,
                    msgctxt=context,
                    obsolete=obsolete,
                )
                entry.msgstr_plural = msgstr_plural
            else:
                # Singular entry
                msgstr = row[3] or "" if len(row) > 3 else ""
                entry = polib.POEntry(
                    msgid=msgid,
                    msgstr=msgstr,
                    msgctxt=context,
                    obsolete=obsolete,
                )

            po.append(entry)

        wb.close()

        # Write the PO file
        out_filename = self.options.filename.format(lang=language)
        out_path = self.options.outdir / out_filename
        po.save(str(out_path))
        self.ok(f"Created {out_path}")


def main() -> None:
    """Run the Excel-to-PO CLI command."""
    parser = argparse.ArgumentParser(
        prog="xop-convert",
        formatter_class=ArgumentFormatter,
        description=dedent(
            """
            Convert Excel Spreadsheets back to .po files.

                %(prog)s path/to/translations.xlsx
            """,
        ),
    )
    parser.add_argument(
        nargs="+",
        type=str,
        dest="xlsx_file",
        metavar="PATH_TO_XLSX_FILE",
        help="Space separated path to one or more .xlsx files.",
    )
    parser.add_argument(
        "-o",
        "--outdir",
        type=Path,
        default=Path(),
        help="The path to store the generated .po files. Default: .",
    )
    parser.add_argument(
        "-f",
        "--filename",
        type=str,
        default="{lang}.po",
        help=dedent(
            """
            The filename format for the generated .po files. You can use
            the variable `{lang}`.

            Default: {lang}.po
            """,
        ),
    )
    options = parser.parse_args()

    convert = Excel2PoConverter(options=options)
    convert.run()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.stdout.write("\nOK, bye \U0001f622\n")
