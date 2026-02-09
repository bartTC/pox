"""Spreadsheet styles for pox."""

from enum import Enum

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.fills import FILL_SOLID

FONT_FAMILY = "Tahoma"
FONT_SIZE = 14

WHITE = "FFFFFFFF"
BLACK = "FF000000"
LIGHT_GRAY = "FFF2F2F2"
MEDIUM_GRAY = "FFD9D9D9"
DARK_GRAY = "FF595959"
YELLOW = "FFFFF2CC"
RED_BG = "FFCC0000"
OBSOLETE_GRAY = "FFBFBFBF"
ROW_STRIPE = "FFF9F9F9"

DEFAULT_ALIGNMENT = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=True,
    shrinkToFit=False,
)

MEDIUM_BLACK = Side(style="medium", color=BLACK)
THIN_BLACK = Side(style="thin", color=BLACK)

EMPTY_BORDER = Border(
    left=MEDIUM_BLACK,
    right=MEDIUM_BLACK,
    bottom=THIN_BLACK,
)

EMPTY_BORDER_TOP = Border(
    left=MEDIUM_BLACK,
    right=MEDIUM_BLACK,
    top=MEDIUM_BLACK,
    bottom=THIN_BLACK,
)

EMPTY_BORDER_BOTTOM = Border(
    left=MEDIUM_BLACK,
    right=MEDIUM_BLACK,
    bottom=MEDIUM_BLACK,
)

EMPTY_BORDER_SINGLE = Border(
    left=MEDIUM_BLACK,
    right=MEDIUM_BLACK,
    top=MEDIUM_BLACK,
    bottom=MEDIUM_BLACK,
)


class SpreadsheetStyles(Enum):
    """Define the openpyxl cell styles used in the spreadsheet."""

    SHEET_HEADLINE = {  # noqa: RUF012
        "fill": PatternFill(fill_type=FILL_SOLID, fgColor=RED_BG),
        "font": Font(
            name=FONT_FAMILY,
            size=22,
            bold=True,
            italic=False,
            color=WHITE,
        ),
        "alignment": Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=False,
            shrinkToFit=False,
        ),
    }

    ID = {  # noqa: RUF012
        "font": Font(name=FONT_FAMILY, size=FONT_SIZE, color=DARK_GRAY),
        "alignment": DEFAULT_ALIGNMENT,
    }

    HEADER = {  # noqa: RUF012
        "fill": PatternFill(fill_type=FILL_SOLID, fgColor=BLACK),
        "font": Font(name=FONT_FAMILY, size=FONT_SIZE, bold=True, color=WHITE),
        "alignment": DEFAULT_ALIGNMENT,
    }

    CONTEXT = {  # noqa: RUF012
        "fill": PatternFill(fill_type=FILL_SOLID, fgColor=LIGHT_GRAY),
        "font": Font(name=FONT_FAMILY, size=FONT_SIZE, italic=True, color=DARK_GRAY),
        "alignment": DEFAULT_ALIGNMENT,
    }

    MSG_ID = {  # noqa: RUF012
        "font": Font(name=FONT_FAMILY, size=FONT_SIZE, color=BLACK),
        "alignment": DEFAULT_ALIGNMENT,
    }

    MSG_STR_EMPTY = {  # noqa: RUF012
        "fill": PatternFill(fill_type=FILL_SOLID, fgColor=YELLOW),
        "font": Font(name=FONT_FAMILY, size=FONT_SIZE, color=BLACK),
        "alignment": DEFAULT_ALIGNMENT,
    }

    MSG_STR_OBSOLETE = {  # noqa: RUF012
        "font": Font(
            name=FONT_FAMILY,
            size=FONT_SIZE,
            italic=True,
            color=OBSOLETE_GRAY,
        ),
        "alignment": DEFAULT_ALIGNMENT,
    }
