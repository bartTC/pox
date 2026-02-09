"""Tests for PO-to-Excel conversion (pox-convert)."""

from pathlib import Path

from openpyxl import load_workbook

from pox.spreadsheet import SpreadsheetGenerator

from .conftest import FIXTURES, build_context


def test_generates_xlsx(tmp_path: Path, singular_po: Path) -> None:
    """Generated xlsx file exists and has the correct suffix."""
    context = build_context(singular_po)
    gen = SpreadsheetGenerator(outdir=tmp_path)
    result = gen.generate(filename="test_{lang}.xlsx", context=context)
    assert result.exists()
    assert result.suffix == ".xlsx"


def test_translations_sheet_exists(singular_xlsx: Path) -> None:
    """Workbook contains a sheet starting with 'Translations'."""
    wb = load_workbook(str(singular_xlsx))
    assert any(s.startswith("Translations") for s in wb.sheetnames)
    wb.close()


def test_header_row(singular_xlsx: Path) -> None:
    """Header row contains expected column names."""
    wb = load_workbook(str(singular_xlsx))
    ws = wb.worksheets[0]
    header = [c.value for c in ws[1]]
    assert "id" in header
    assert "Context" in header
    assert "Singular Form" in header
    assert any("Translation" in str(h) for h in header)
    wb.close()


def test_header_is_bold(singular_xlsx: Path) -> None:
    """Header cells use bold font."""
    wb = load_workbook(str(singular_xlsx))
    ws = wb.worksheets[0]
    header_cell = ws.cell(row=1, column=3)
    assert header_cell.font.bold is True
    wb.close()


def test_header_row_is_frozen(singular_xlsx: Path) -> None:
    """First row is frozen for scrolling."""
    wb = load_workbook(str(singular_xlsx))
    ws = wb.worksheets[0]
    assert ws.freeze_panes == "A2"
    wb.close()


def test_singular_entries(singular_xlsx: Path) -> None:
    """Singular message IDs appear in the spreadsheet."""
    wb = load_workbook(str(singular_xlsx))
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    msgids = [r[2] for r in rows[1:] if r[0]]
    assert "Hello" in msgids
    assert "Goodbye" in msgids
    assert "Open" in msgids
    wb.close()


def test_empty_translation_has_yellow_fill(singular_xlsx: Path) -> None:
    """Empty translation cells have yellow fill."""
    wb = load_workbook(str(singular_xlsx))
    ws = wb.worksheets[0]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[2] == "Goodbye":
            cell = ws.cell(row=row_idx, column=4)
            assert cell.fill.fgColor.rgb == "FFFFF2CC"
            break
    wb.close()


def test_empty_translation_has_border(singular_xlsx: Path) -> None:
    """Empty translation cells have left and right borders."""
    wb = load_workbook(str(singular_xlsx))
    ws = wb.worksheets[0]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[2] == "Goodbye":
            cell = ws.cell(row=row_idx, column=4)
            assert cell.border.left.style is not None
            assert cell.border.right.style is not None
            break
    wb.close()


def test_context_column(singular_xlsx: Path) -> None:
    """Context column contains the correct msgctxt value."""
    wb = load_workbook(str(singular_xlsx))
    ws = wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        if row[2] == "Open":
            assert row[1] == "adjective"
            break
    wb.close()


def test_custom_properties(singular_xlsx: Path) -> None:
    """Custom document property Language is set correctly."""
    wb = load_workbook(str(singular_xlsx))
    lang_prop = None
    for prop in wb.custom_doc_props:
        if prop.name == "Language":
            lang_prop = prop.value
    assert lang_prop == "de"
    wb.close()


def test_plural_columns(plurals_xlsx: Path) -> None:
    """Plural entries produce two translation columns."""
    wb = load_workbook(str(plurals_xlsx))
    ws = wb.worksheets[0]
    header = [c.value for c in ws[1]]
    translation_headers = [h for h in header if h and "Translation" in str(h)]
    assert len(translation_headers) == 2
    wb.close()


def test_plural_entry_values(plurals_xlsx: Path) -> None:
    """Plural translation values are written correctly."""
    wb = load_workbook(str(plurals_xlsx))
    ws = wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        if row[2] == "%d items":
            assert row[3] == "%d Gegenstand"
            assert row[4] == "%d Gegenstaende"
            break
    wb.close()


def test_gridlines_disabled(singular_xlsx: Path) -> None:
    """Sheet gridlines are turned off."""
    wb = load_workbook(str(singular_xlsx))
    ws = wb.worksheets[0]
    assert ws.sheet_view.showGridLines is False
    wb.close()


def test_obsolete_entry(tmp_path: Path, obsolete_po: Path) -> None:
    """Obsolete entries are marked with 'obsolete' context."""
    context = build_context(obsolete_po)
    gen = SpreadsheetGenerator(outdir=tmp_path)
    result = gen.generate(filename="test_{lang}.xlsx", context=context)
    wb = load_workbook(str(result))
    ws = wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        if row[2] == "Old":
            assert row[1] == "obsolete"
            break
    wb.close()


def test_fuzzy_entry_included(tmp_path: Path, fuzzy_po: Path) -> None:
    """Fuzzy entries are included when build_context doesn't filter them."""
    context = build_context(fuzzy_po)
    gen = SpreadsheetGenerator(outdir=tmp_path)
    result = gen.generate(filename="test_{lang}.xlsx", context=context)
    wb = load_workbook(str(result))
    ws = wb.worksheets[0]
    msgids = [r[2] for r in ws.iter_rows(values_only=True) if r[0]]
    # build_context includes all entries (no fuzzy filtering)
    assert "Maybe" in msgids
    wb.close()


def test_no_style_cells_handled(tmp_path: Path, singular_po: Path) -> None:
    """Cells with no style tuple (None) are handled without error."""
    context = build_context(singular_po)
    gen = SpreadsheetGenerator(outdir=tmp_path)
    result = gen.generate(filename="test_{lang}.xlsx", context=context)
    assert result.exists()


def test_non_contiguous_empty_borders(tmp_path: Path, mixed_empty_po: Path) -> None:
    """Non-contiguous empty rows get separate border groups."""
    context = build_context(mixed_empty_po)
    gen = SpreadsheetGenerator(outdir=tmp_path)
    result = gen.generate(filename="test_{lang}.xlsx", context=context)
    wb = load_workbook(str(result))
    ws = wb.worksheets[0]

    # "Goodbye" (row 3) is a single empty, should have top+bottom border
    cell_goodbye = ws.cell(row=3, column=4)
    assert cell_goodbye.border.top.style is not None
    assert cell_goodbye.border.bottom.style is not None

    # "Foo" and "Bar" (rows 5-6) are contiguous empties
    cell_foo = ws.cell(row=5, column=4)
    assert cell_foo.border.top.style is not None  # top of group

    cell_bar = ws.cell(row=6, column=4)
    assert cell_bar.border.bottom.style is not None  # bottom of group
    wb.close()


def test_empty_plural_form_highlighted(tmp_path: Path) -> None:
    """A plural entry with one empty form gets yellow fill."""
    context = build_context(FIXTURES / "plural_empty.po")
    gen = SpreadsheetGenerator(outdir=tmp_path)
    result = gen.generate(filename="test_{lang}.xlsx", context=context)
    wb = load_workbook(str(result))
    ws = wb.worksheets[0]
    # Row 2 is the plural entry, column 5 (plural form 2) should be yellow
    cell = ws.cell(row=2, column=5)
    assert cell.fill.fgColor.rgb == "FFFFF2CC"
    wb.close()


def test_three_plural_columns(tmp_path: Path, plurals_polish_po: Path) -> None:
    """Polish has 3 plural forms, should produce 3 translation columns."""
    context = build_context(plurals_polish_po)
    gen = SpreadsheetGenerator(outdir=tmp_path)
    result = gen.generate(filename="test_{lang}.xlsx", context=context)
    wb = load_workbook(str(result))
    ws = wb.worksheets[0]
    header = [c.value for c in ws[1]]
    translation_headers = [h for h in header if h and "Translation" in str(h)]
    assert len(translation_headers) == 3
    wb.close()


def test_three_plural_values(tmp_path: Path, plurals_polish_po: Path) -> None:
    """All 3 Polish plural forms are written to the spreadsheet."""
    context = build_context(plurals_polish_po)
    gen = SpreadsheetGenerator(outdir=tmp_path)
    result = gen.generate(filename="test_{lang}.xlsx", context=context)
    wb = load_workbook(str(result))
    ws = wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        if row[2] == "%d items":
            assert row[3] == "%d element"
            assert row[4] == "%d elementy"
            assert row[5] == "%d elementow"
            break
    wb.close()


def test_six_plural_columns_arabic(tmp_path: Path) -> None:
    """Arabic has 6 plural forms, should produce 6 translation columns."""
    context = build_context(FIXTURES / "plurals_arabic.po")
    gen = SpreadsheetGenerator(outdir=tmp_path)
    result = gen.generate(filename="test_{lang}.xlsx", context=context)
    wb = load_workbook(str(result))
    ws = wb.worksheets[0]
    header = [c.value for c in ws[1]]
    translation_headers = [h for h in header if h and "Translation" in str(h)]
    assert len(translation_headers) == 6

    # Verify all 6 forms are present in the data row
    for row in ws.iter_rows(values_only=True):
        if row[2] == "%d items":
            for col in range(3, 9):
                assert row[col] is not None
                assert row[col] != ""
            break
    wb.close()


def test_no_language_po_defaults_to_en(tmp_path: Path, no_language_po: Path) -> None:
    """PO file without Language metadata defaults to 'en'."""
    context = build_context(no_language_po)
    assert context.language == "en"
    gen = SpreadsheetGenerator(outdir=tmp_path)
    result = gen.generate(filename="test_{lang}.xlsx", context=context)
    assert result.exists()
