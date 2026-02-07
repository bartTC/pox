import argparse
from datetime import UTC, datetime
from pathlib import Path

import polib
from openpyxl import load_workbook

from pox.cli.xop_convert import Excel2PoConverter
from pox.datastructures import (
    Message,
    PluralTranslation,
    SingularTranslation,
    SpreadsheetContext,
)
from pox.plurals import get_plural_hints
from pox.spreadsheet import SpreadsheetGenerator


def make_po_file(tmp_path: Path) -> Path:
    """Create a sample PO file with singular, plural, fuzzy, and obsolete entries."""
    po = polib.POFile()
    po.metadata = {
        "Content-Type": "text/plain; charset=UTF-8",
        "Language": "de",
        "Plural-Forms": "nplurals=2; plural=(n != 1);",
    }

    # Singular entry
    po.append(
        polib.POEntry(
            msgid="Hello",
            msgstr="Hallo",
        ),
    )

    # Empty singular entry
    po.append(
        polib.POEntry(
            msgid="Goodbye",
            msgstr="",
        ),
    )

    # Entry with context
    po.append(
        polib.POEntry(
            msgid="Open",
            msgstr="Offen",
            msgctxt="adjective",
        ),
    )

    # Plural entry
    entry = polib.POEntry(
        msgid="%d item",
        msgid_plural="%d items",
    )
    entry.msgstr_plural = {0: "%d Gegenstand", 1: "%d Gegenstaende"}
    po.append(entry)

    # Fuzzy entry
    fuzzy = polib.POEntry(
        msgid="Maybe",
        msgstr="Vielleicht",
    )
    fuzzy.flags = ["fuzzy"]
    po.append(fuzzy)

    # Obsolete entry
    po.append(
        polib.POEntry(
            msgid="Old",
            msgstr="Alt",
            obsolete=True,
        ),
    )

    path = tmp_path / "test.po"
    po.save(str(path))
    return path


def build_context_from_po(po_path: Path) -> SpreadsheetContext:
    """Build a SpreadsheetContext from a PO file, mimicking pox_convert logic."""
    pofile = polib.pofile(str(po_path))
    messages = []

    for item in pofile:
        if item.msgstr_plural:
            t = PluralTranslation(
                msgid=item.msgid_plural,
                msgstr=item.msgstr_plural,
            )
        else:
            t = SingularTranslation(
                msgid=item.msgid,
                msgstr=item.msgstr,
            )

        messages.append(
            Message(
                translation=t,
                context=item.msgctxt,
                comment=item.comment,
                tcomment=item.tcomment,
                obsolete=item.obsolete == 1,
            ),
        )

    return SpreadsheetContext(
        created=datetime.now(tz=UTC),
        messages=messages,
        language=pofile.metadata.get("Language", "en"),
        plural_form_hints=get_plural_hints(pofile.metadata.get("Plural-Forms")),
    )


class TestPluralHints:
    def test_no_plural_form(self):
        result = get_plural_hints(None)
        assert result == {0: "Singular", 1: "Plural"}

    def test_two_forms(self):
        result = get_plural_hints("nplurals=2; plural=(n != 1);")
        assert len(result) == 2

    def test_three_forms(self):
        result = get_plural_hints(
            "nplurals=3; plural=(n%10==1 && n%100!=11 ? 0 : n%10>=2 && n%10<=4 && (n%100<10 || n%100>=20) ? 1 : 2);",
        )
        assert len(result) == 3


class TestSpreadsheetGeneration:
    def test_generates_xlsx(self, tmp_path):
        po_path = make_po_file(tmp_path)
        context = build_context_from_po(po_path)
        gen = SpreadsheetGenerator(outdir=tmp_path)
        result = gen.generate(filename="test_{lang}.xlsx", context=context)
        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_xlsx_has_translations_sheet(self, tmp_path):
        po_path = make_po_file(tmp_path)
        context = build_context_from_po(po_path)
        gen = SpreadsheetGenerator(outdir=tmp_path)
        result = gen.generate(filename="test_{lang}.xlsx", context=context)
        wb = load_workbook(str(result))
        sheet_names = wb.sheetnames
        assert any(s.startswith("Translations") for s in sheet_names)
        wb.close()

    def test_singular_entries_in_xlsx(self, tmp_path):
        po_path = make_po_file(tmp_path)
        context = build_context_from_po(po_path)
        gen = SpreadsheetGenerator(outdir=tmp_path)
        result = gen.generate(filename="test_{lang}.xlsx", context=context)
        wb = load_workbook(str(result))
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        # Row 0 = fill hint, Row 1 = header, Row 2+ = data
        msgids = [r[2] for r in rows[2:-1] if r[0]]
        assert "Hello" in msgids
        assert "Goodbye" in msgids
        wb.close()

    def test_styles_applied(self, tmp_path):
        po_path = make_po_file(tmp_path)
        context = build_context_from_po(po_path)
        gen = SpreadsheetGenerator(outdir=tmp_path)
        result = gen.generate(filename="test_{lang}.xlsx", context=context)
        wb = load_workbook(str(result))
        ws = wb.worksheets[0]
        # Header row (row 2) should have bold font
        header_cell = ws.cell(row=2, column=3)
        assert header_cell.font.bold is True
        wb.close()

    def test_custom_properties(self, tmp_path):
        po_path = make_po_file(tmp_path)
        context = build_context_from_po(po_path)
        gen = SpreadsheetGenerator(outdir=tmp_path)
        result = gen.generate(filename="test_{lang}.xlsx", context=context)
        wb = load_workbook(str(result))
        lang_prop = None
        for prop in wb.custom_doc_props:
            if prop.name == "Language":
                lang_prop = prop.value
        assert lang_prop == "de"
        wb.close()


class TestRoundtrip:
    def test_singular_roundtrip(self, tmp_path):
        """PO -> Excel -> PO preserves singular entries."""
        po_path = make_po_file(tmp_path)
        context = build_context_from_po(po_path)

        gen = SpreadsheetGenerator(outdir=tmp_path)
        xlsx_path = gen.generate(filename="test_{lang}.xlsx", context=context)

        opts = argparse.Namespace(
            xlsx_file=[str(xlsx_path)],
            outdir=tmp_path,
            filename="{lang}_roundtrip.po",
        )
        converter = Excel2PoConverter(options=opts)
        converter.run()

        # Read back the PO
        result_po = polib.pofile(str(tmp_path / "de_roundtrip.po"))
        result_msgids = [e.msgid for e in result_po]

        assert "Hello" in result_msgids
        assert "Goodbye" in result_msgids

    def test_metadata_preserved(self, tmp_path):
        """Language metadata survives roundtrip."""
        po_path = make_po_file(tmp_path)
        context = build_context_from_po(po_path)
        gen = SpreadsheetGenerator(outdir=tmp_path)
        xlsx_path = gen.generate(filename="test_{lang}.xlsx", context=context)

        opts = argparse.Namespace(
            xlsx_file=[str(xlsx_path)],
            outdir=tmp_path,
            filename="{lang}_rt.po",
        )
        converter = Excel2PoConverter(options=opts)
        converter.run()

        result_po = polib.pofile(str(tmp_path / "de_rt.po"))
        assert result_po.metadata["Language"] == "de"
